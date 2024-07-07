import os
import re
import requests as req
import codecs
import json
import time
from datetime import datetime
from bs4 import BeautifulSoup
import random
import platform
import logging

# Настройка логгирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

session = req.Session()

headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Accept-Encoding': 'gzip, deflate, br',
    'Accept-Language': 'en-US,en;q=0.9',
    'Cache-Control': 'no-cache',
    'Cookie': '',
    'Pragma': 'no-cache',
    'Sec-Ch-Ua': '"Not.A/Brand";v="8", "Chromium";v="114"',
    'Sec-Ch-Ua-Mobile': '?0',
    'Sec-Ch-Ua-Platform': '"Windows"',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'none',
    'Sec-Fetch-User': '?1',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36'
}

urls = {
    'https://urfu.ru/ru/ratings/': 'rating',
    'https://urfu.ru/ru/ratings-contract/': 'contracts',
    'https://urfu.ru/ru/ratings-today/': 'ratings-today',
    'https://urfu.ru/ru/ratings-masters/': 'ratings-masters',
    'https://urfu.ru/ru/ratings-masters-full/': 'ratings-masters-full',
    'https://urfu.ru/ru/ratings-masters-extramural/': 'ratings-masters-extramural'
}
temp_path = "R:\\"
main_dir = os.path.dirname(os.path.abspath(__file__))
css_path = os.path.join(main_dir, "entrants.css")
files_non_person_data_path = os.path.join(main_dir, "Окончательные_файлы_без_персон_данных\\")
files_path = os.path.join(main_dir, "Окончательные_файлы\\")
#sec = 1800 #30 минут

def ratingtodaycheck(url, default_str):
    return urls.get(url, default_str)

def create_directory(path):
    if not os.path.exists(path):
        os.makedirs(path)

if __name__ == '__main__':
    if platform.system() == 'Windows':
        os.system('chcp 65001')
        os.system('cls')
    
    create_directory(files_path)
    create_directory(files_non_person_data_path)
    
    import jpype
    import asposecells

    jpype.startJVM()
    from asposecells.api import Workbook

    while True:
        for url in urls:
            try:
                requests_urfu = session.get(url, headers=headers)
                if requests_urfu.status_code == 502:
                    print("[!] Умер frontend урфу")
                elif requests_urfu.status_code == 404:
                    print("[!] В данный момент нету результатов или их удалили))")
            except:
                print("[!] Здох сам урфу")
            try:
                soup = BeautifulSoup(requests_urfu.text, "html.parser")
                id = soup.find("div", attrs={"data-plugin": "entrants"}).attrs['data-preset']
                print("[*] Сайт ворк (" + url + ")")
            except:
                print("[!] Сайт не ворк (" + url + ")")
                if url == 'https://urfu.ru/ru/ratings-today/':
                    print("[*] Ок скачаем через апи со стачиским id")
                    id = "56"
                elif url == 'https://urfu.ru/ru/ratings/':
                    print("[*] Ок скачаем через апи со стачиским id")
                    id = "55"
                elif url == 'https://urfu.ru/ru/ratings-contract/':
                    print("[*] Ок скачаем через апи со стачиским id")
                    id = "48"
                elif url == 'https://urfu.ru/ru/ratings-masters/':
                    print("[*] Ок скачаем через апи со стачиским id")
                    id = "57"
                elif url == 'https://urfu.ru/ru/ratings-masters-full/':
                    print("[*] Ок скачаем через апи со стачиским id")
                    id = "59"
                elif url == 'https://urfu.ru/ru/ratings-masters-extramural/':
                    print("[*] Ок скачаем через апи со стачиским id")
                    id = "58"
                else:
                    continue
            #url_css = soup.find("div", attrs={"data-plugin": "entrants"}).find('link').attrs['href']
            url_css = "https://urfu.ru/typo3conf/ext/urfu_entrants/Resources/Public/css/entrants.css"
            try:
                requests_urfu_json = session.get("https://urfu.ru/api/ratings/info/" + id, headers=headers)
            except:
                print("[!] Умер backend урфу api")
                continue
            requests_urfu_json.encoding = "utf-8-sig"
            try:
                parameters_urfu = json.loads(requests_urfu_json.text)
            except:
                continue
            # print(parameters_urfu)
            for full_vyz in parameters_urfu['institutes']:
                for forms_vyz in full_vyz['eduForms']:
                    try:
                        url_vyz = json.loads(session.get(requests_urfu_json.url + str(forms_vyz['rowId']) + "/", headers=headers, timeout=7).text)
                    except:
                        print("[!] Умер backend урфу api отвечающий за " + full_vyz['title']+ ' ' + forms_vyz['title'])
                        continue
                    # for elem in updatedAt:
                    # print(str(parameters["updatedAt"]))
                    # if url_vyz["updatedAt"] != elem:
                    # updatedAt.append(url_vyz["updatedAt"])
                    # updatedAt[index] = url_vyz["updatedAt"]
                    # print(updatedAt)
                    date_object = datetime.strptime(url_vyz["updatedAt"], "%d.%m.%Y %H:%M:%S")
                    updatedAt_filename = str(date_object.date().strftime("%d-%m-%Y")) + "__" + str(
                        date_object.time().strftime("%H-%M-%S"))
                    folder_date = str(date_object.date().strftime("%d-%m-%Y"))
                    #create_directory(files_path + ratingtodaycheck(url, parameters_urfu['type']))
                    create_directory(files_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date)
                    #create_directory(files_non_person_data_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date)
                    create_directory(files_non_person_data_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date)
                    #try:
                    #    files = os.listdir(files_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date)
                    #except:
                     #   os.mkdir(files_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date)
                     #   files = os.listdir(files_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date)
                    isExist=False
                    #for f in files:
                        #if f == full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + ".xlsx":
                            #isExist=True
                    isExist=(os.path.exists(files_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\EXCEL\\" + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + ".xlsx") and \
                    os.path.exists(files_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\HTML\\" + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + ".html") and \
                    os.path.exists(files_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\PDF\\" + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + ".pdf") and \
                    os.path.exists(files_non_person_data_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\EXCEL\\" + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + ".xlsx") and \
                    os.path.exists(files_non_person_data_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\HTML\\" + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + ".html") and \
                    os.path.exists(files_non_person_data_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\PDF\\" + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + ".pdf"))
                    if not isExist:
                        #print("[+] Загрузка html-файла (" + ratingtodaycheck(url, parameters_urfu['type']) + ", id=" + id +"): " + full_vyz['title'] + ' ' + forms_vyz['title'] + "___" + str(forms_vyz['rowId']) + "____" + url_vyz["updatedAt"])
                        try:
                            fp = session.get("https://urfu.ru" + url_vyz['url'], headers=headers, timeout=7)
                            if (fp.status_code == 502):
                                print("[!] Не удалось скачать html-файл (" + ratingtodaycheck(url, parameters_urfu[
                                    'type']) + ", id=" + id + "): " + full_vyz['title'] + ' ' + forms_vyz[
                                          'title'] + "___" + str(
                                    forms_vyz['rowId']) + "____" + url_vyz["updatedAt"])
                                continue
                            else:
                                print("[+] Загрузка html-файла (" + ratingtodaycheck(url, parameters_urfu[
                                    'type']) + ", id=" + id + "): " + full_vyz['title'] + ' ' + forms_vyz[
                                          'title'] + "___" + str(
                                    forms_vyz['rowId']) + "____" + url_vyz["updatedAt"])
                        except:
                            print("[!] Не удалось скачать html-файл (" + ratingtodaycheck(url, parameters_urfu[
                                'type']) + ", id=" + id + "): " + full_vyz['title'] + ' ' + forms_vyz[
                                      'title'] + "___" + str(
                                forms_vyz['rowId']) + "____" + url_vyz["updatedAt"])
                            continue
                        fp.encoding = "utf-8-sig"
                        soup_html_table = BeautifulSoup(fp.text, "html.parser")
                        link = soup_html_table.new_tag('link')
                        link.attrs['rel'] = 'stylesheet'
                        link.attrs['href'] = "entrants.css"
                        # <link rel="stylesheet" href="https://urfu.ru/typo3conf/ext/urfu_entrants/Resources/Public/css/entrants.css">
                        soup_html_table.contents[-1].insert_after(link)
                        soup_html_table_itog = BeautifulSoup("<html>"
                                                             "<body>"
                                                             "<div data-plugin=\"entrants\" data-preset=\"" + id + "\" class=\"urfu-entrants-wrapper\">"
                                                             "<div class=\"rating-container\" style=\"display: block;\">"
                                                             "</div>"
                                                             "</div>"
                                                             "</body>"
                                                             "</html>", "html.parser")
                        div1 = soup_html_table_itog.find('div', class_='urfu-entrants-wrapper')
                        div2 = soup_html_table_itog.find('div', class_='rating-container')
                        div2.append(soup_html_table.prettify(formatter=None))

                        # Находим последний элемент body
                        # last_element = soup_html_table.contents[]

                        ## Устанавливаем атрибут class для div1
                        # div1.attrs['class'] = 'urfu-entrant-wrapper'

                        ## Устанавливаем атрибуты class и style для div2
                        # div2.attrs['class'] = 'rating-container'
                        # div2.attrs['style'] = 'display: block;'

                        # Вставляем div2 перед последним элементом body
                        # soup_html_table_itog.contents[-1].insert_before(div2)
                        # Перемещаем последний элемент body в div2
                        # div2.append(last_element)
                        # Вставляем div1 перед div2
                        # div2.insert_before(div1)

                        # Выводим измененный HTML-код
                        #print(soup_html_table_itog.prettify(formatter=None))
                        create_directory(files_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\HTML")
                        create_directory(files_non_person_data_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\HTML")
                        if not os.path.exists(files_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\HTML\\entrants.css") and not os.path.exists(files_non_person_data_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\HTML\\entrants.css"):
                            try:
                                css = session.get(url_css, headers=headers, timeout=10).text
                                if (css.status_code == 502):
                                    os.system('copy ' + css_path + " " + files_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\HTML\\entrants.css")
                                    os.system('copy ' + css_path + " " + files_non_person_data_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\HTML\\entrants.css")
                                else:
                                    file = codecs.open(files_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\HTML\\entrants.css", "w", "utf-8")
                                    file.write(css)
                                    file.close()
                                    file = codecs.open(files_non_person_data_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\HTML\\entrants.css", "w", "utf-8")
                                    file.write(css)
                                    file.close()
                            except:
                                os.system('copy ' + css_path + " " + files_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\HTML\\entrants.css")
                                os.system('copy ' + css_path + " " + files_non_person_data_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\HTML\\entrants.css")
                        file = codecs.open(files_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\HTML\\" + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + ".html", "w", "utf-8")
                        file.write(soup_html_table_itog.prettify(formatter=None))
                        text_html = soup_html_table_itog.prettify(formatter=None)
                        file.close()
                        masked_html = re.sub(r'\b\d{11}\b', '***********', text_html)
                        file = codecs.open(files_non_person_data_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\HTML\\" + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + ".html", "w", "utf-8")
                        file.write(masked_html)
                        file.close()
                        file = codecs.open(temp_path + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + ".html", "w", "utf-8")
                        file.write(soup_html_table_itog.prettify(formatter=None))
                        file.close()
                        create_directory(files_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\PDF")
                        create_directory(files_non_person_data_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\PDF")

                        import pdfkit
                        path_wkhtmltopdf = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'
                        __config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)
                        options = {
                                    'page-size': 'A3',
                                    'margin-top': '10mm',
                                    'margin-right': '10mm',
                                    'margin-bottom': '10mm',
                                    'margin-left': '10mm',
                                    'encoding': 'UTF-8'#,
                        }
                        try:
                            pdfkit.from_string(text_html, files_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\PDF\\" + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + ".pdf",css=css_path, configuration=__config, options=options)
                        except:
                            None
                        try:
                            pdfkit.from_string(masked_html, files_non_person_data_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\PDF\\" + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + ".pdf",css=css_path, configuration=__config, options=options)
                        except:
                            None
                        #os.add_dll_directory(r"C:\Program Files\GTK3-Runtime Win64\bin")
                        #HTML().write_pdf()
                        # from asposecells.api import HtmlLoadOptions

                        create_directory(files_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\EXCEL")
                        create_directory(files_non_person_data_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\EXCEL")

                        workbook = Workbook(temp_path + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + ".html")
                        workbook_masked = Workbook(files_non_person_data_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\HTML\\" + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + ".html")
                        workbook.save(temp_path + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + "_notedit.xlsx")
                        workbook_masked.save(temp_path + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + "_notedit_masked.xlsx")
                        from openpyxl import load_workbook

                        # Загрузить файл Excel
                        excel_file = temp_path + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + "_notedit.xlsx"
                        wb = load_workbook(excel_file)
                        # Выбрать лист, который нужно удалить (например, 'Лист1')
                        sheet_to_delete = 'Evaluation Warning'
                        # Удалить выбранный лист
                        if sheet_to_delete in wb.sheetnames:
                            sheet = wb[sheet_to_delete]
                            wb.remove(sheet)
                        # Сохранить изменения в файле
                        wb.save(files_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\EXCEL\\" + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + ".xlsx")
                        wb2 = load_workbook(temp_path + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + "_notedit_masked.xlsx")
                        # Удалить выбранный лист
                        if sheet_to_delete in wb2.sheetnames:
                            sheet = wb2[sheet_to_delete]
                            wb2.remove(sheet)
                        # Сохранить изменения в файле
                        wb.save(files_non_person_data_path + ratingtodaycheck(url, parameters_urfu['type']) + "\\" + folder_date + "\\EXCEL\\" + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + ".xlsx")
                        os.remove(temp_path + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename +".html")
                        os.remove(temp_path + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + "_notedit.xlsx")
                        os.remove(temp_path + full_vyz['title'] + "__" + forms_vyz['title'] + "__" + updatedAt_filename + "_notedit_masked.xlsx")
                    else:
                        print("[+] Файл есть (" + ratingtodaycheck(url, parameters_urfu['type']) + ", id=" + id +"): " + full_vyz['title'] + ' ' + forms_vyz['title'] + "___" + str(forms_vyz['rowId']) + "____" + url_vyz["updatedAt"])
        secs = random.randint(1800, 10800)
        print('[!] restart через ' + str(secs))
        time.sleep(secs)
        os.system('cls')
